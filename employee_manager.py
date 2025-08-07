import tkinter as tk
import openpyxl
import os.path
from tkinter import messagebox


class EmployeeManager:

    def __init__(self, master):


        self.master = master
        self.master.title("مدیریت کارکنان")
        self.master.geometry('800x600')

        self.filepath = 'employees.xlsx'


        input_frame = tk.Frame(self.master, pady=10)
        input_frame.pack(fill='x', padx=10)



        # فیلد کارمندی
        lbl_id = tk.Label(input_frame, text='کد کارمندی')
        lbl_id.pack(side='right')
        self.ent_id = tk.Entry(input_frame)
        self.ent_id.pack(side='right')

        # فیلد نام و نام خانوادگی
        lbl_name = tk.Label(input_frame, text='نام و نام خانوادگی')
        lbl_name.pack(side='right')
        self.ent_name = tk.Entry(input_frame)
        self.ent_name.pack(side='right')

        # فیلد حقوق دریافتی
        lbl_salary = tk.Label(input_frame, text='حقوق دریافتی')
        lbl_salary.pack(side='right')
        self.ent_salary = tk.Entry(input_frame)
        self.ent_salary.pack(side='right')


        # دکمه ذخیره
        btn_save = tk.Button(input_frame, text='ذخیره کارمند جدید', command=self.save_employee_data)
        btn_save.pack(pady=10, side='left')


        # --- فریم نمایش لیست (وسط) ---
        display_frame = tk.Frame(self.master, pady=10)
        display_frame.pack(expand=True, fill='both', padx=10)

        # ساخت نوار اسکرول
        scrollbar = tk.Scrollbar(display_frame)
        scrollbar.pack(side='right', fill='y')

        # ساخت لیست‌باکس و اتصال آن به اسکرول‌بار
        self.listbox = tk.Listbox(display_frame, yscrollcommand=scrollbar.set)
        self.listbox.pack(side='left', expand=True, fill='both')
        # اتصال اسکرول‌بار به لیست‌باکس
        scrollbar.config(command=self.listbox.yview)


        # --- فریم دکمه‌های عملیاتی (پایین) ---
        action_frame = tk.Frame(self.master, pady=10)
        action_frame.pack(fill='x', padx=10)



        # متغیر برای نگهداری انتخاب کاربر (جستجو با کد یا نام)
        self.search_criteria = tk.StringVar(value='id')

        radio_id= tk.Radiobutton(action_frame, text='کد', variable=self.search_criteria, value='id')
        radio_id.pack(side='right', padx=5)

        radio_name= tk.Radiobutton(action_frame, text='نام و نام خانوادگی', variable=self.search_criteria, value='name')
        radio_name.pack(side='right', padx=5)



        
        lbl_search = tk.Label(action_frame, text='جستجو')
        lbl_search.pack(side='right', padx=5)
        self.ent_search = tk.Entry(action_frame)
        self.ent_search.pack(side='right', fill='x', padx=5)


        btn_search = tk.Button(action_frame, text='جستجو کن', command=self.search_employees)
        btn_search.pack(side='right')
        
        


        btn_display = tk.Button(action_frame, text='نمایش همه کارکنان', command=self.display_employees)
        btn_display.pack(side='left')


    def save_employee_data(self):


        employee_id = self.ent_id.get()
        employee_name = self.ent_name.get()
        salary = self.ent_salary.get()


        if not employee_id or not employee_name or not salary:
            messagebox.showerror('لطفا تمام فیلد ها را پر کنید','خطا')
            return
        
        record = [employee_id,employee_name,salary]

        if not os.path.exists(self.filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['کد کارمندی','نام خانوادگی','حقوق دریافتی'])
        else:
            workbook = openpyxl.load_workbook(self.filepath)
            sheet = workbook.active

        sheet.append(record)
        workbook.save(self.filepath)


        self.ent_id.delete(0, tk.END)
        self.ent_name.delete(0, tk.END)
        self.ent_salary.delete(0, tk.END)
        self.ent_id.focus_set()

        self.display_employees()


    # برای نمایش اطلاعات
    def display_employees(self):

        
        self.listbox.delete(0, tk.END)


        if not os.path.exists(self.filepath):
            self.listbox.insert(tk.END, 'فایلی برای نمایش وجود ندارد!')
            return

        workbook = openpyxl.load_workbook(self.filepath)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only = True):

            if row[0] is None:
                continue

            display_text = f"کد: {row[0]} | نام و نام خانوادگی: {row[1]} | حقوق: {row[2]}"
            self.listbox.insert(tk.END, f"{display_text:>140}")





    def search_employees(self):


        # خواندن اطلاعات از رابط کاربری
        search_term = self.ent_search.get().lower()
        criteria = self.search_criteria.get()

        if not search_term:
            return
        

        self.listbox.delete(0, tk.END)

        if not os.path.exists(self.filepath):
            self.listbox.insert(tk.END, 'فایل داده وجود ندارد.')
            return
        
        workbook = openpyxl.load_workbook(self.filepath)
        sheet = workbook.active


        found_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            employee_id = str(row[0]).lower()
            employee_name = str(row[1]).lower()

            match = False

            if criteria == 'id':
                if search_term in employee_id:
                    match = True
            elif criteria == 'name':
                if search_term in employee_name:
                    match = True

            if match:
                display_text = f"کد: {row[0]} | نام: {row[1]} | حقوق: {row[2]}"
                self.listbox.insert(tk.END, f"{display_text:>140}")
                found_count += 1

        if found_count == 0:
            self.listbox.insert(tk.END, 'هیچ نتیجه ای یافت نشد')

            









if __name__ == '__main__':

    root = tk.Tk()

    app = EmployeeManager(root)

    root.mainloop()



